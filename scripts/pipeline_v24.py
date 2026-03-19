#!/usr/bin/env python3
"""
BIST Bilanço Sonrası Ucuzlama Tarayıcısı v2.4 — Deterministik Pipeline
========================================================================
Girdi:  master.tsv + hist_master.tsv (Fintables MCP'den çekilmiş)
Çıktı:  BIST_Qx_YYYY_6YONTEM_TARAMA.xlsx (renkli, 2 sayfa)

Kullanım:
  python pipeline_v24.py master.tsv hist_master.tsv [--faiz 42.5] [--donem Q4_2025]
  python pipeline_v24.py --help

Aynı TSV girdisi → her zaman aynı Excel çıktısı (deterministik).
"""

import sys
import csv
import math
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ openpyxl yüklü değil. pip install openpyxl")


# ─── TCMB FAİZ → DİNAMİK ÇARPAN ──────────────────────────────────────

def adil_fk(faiz_pct: float, faiz_trendi: str = "sabit") -> float:
    """TCMB politika faizi → Dinamik Adil F/K"""
    risk_primi = 0.05
    adil = 1.0 / (faiz_pct / 100.0 * 0.70 + risk_primi)
    if faiz_trendi == "indirim":
        adil *= 1.20
    elif faiz_trendi == "artirim":
        adil *= 0.90
    return round(adil, 2)


SEKTOR_CARPAN = {
    "Sanayi": 1.00, "Üretim": 1.00, "İmalat": 1.00,
    "Teknoloji": 1.875, "Bilişim": 1.875, "Yazılım": 1.875, "Bilişim ve Yazılım": 1.875,
    "Savunma": 1.875, "Havacılık": 1.875,
    "Bankacılık": 0.50, "Banka": 0.50,
    "Enerji": 0.75,
    "Perakende": 1.00,
    "İnşaat": 0.625,
    "Turizm": 1.00,
    "Telekom": 0.875,
    "Tarım": 1.00, "Gıda": 1.00, "Gıda & İçecek": 1.00,
    "Sigorta": 0.75,
    "Maden": 0.75, "Metal": 0.75, "Kimya": 0.85,
    "Gayrimenkul": None,  # NAV bazlı
    "Holding": None,
    "Aracı Kurum": 0.75, "Aracı Kurumlar": 0.75,
    "Faktoring": 0.75, "Leasing": 0.75,
}


def sektor_adil_fk(baz_adil: float, sektor: str) -> float | None:
    carpan = SEKTOR_CARPAN.get(sektor, 1.00)
    if carpan is None:
        return None  # GYO/Holding → NAV
    return round(baz_adil * carpan, 2)


# ─── YARDIMCI FONKSİYONLAR ──────────────────────────────────────────

def safe_float(val, default=None):
    if val is None or val == "" or val == "N/A" or val == "None":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def pct_change(new, old):
    if old is None or new is None or old == 0:
        return None
    return round((new - old) / abs(old) * 100, 1)


def iskonto_pct(mevcut, ortalama):
    if mevcut is None or ortalama is None or ortalama <= 0 or mevcut <= 0:
        return None
    return round((1 - mevcut / ortalama) * 100, 1)


# ─── 6 YÖNTEM SKORLAMA ──────────────────────────────────────────────

def y1_fd_favok_skor(fd_favok_isk):
    """Y1: FD/FAVÖK İskontosu"""
    if fd_favok_isk is None: return 0
    if fd_favok_isk > 50: return 10
    if fd_favok_isk > 30: return 7
    if fd_favok_isk > 15: return 4
    return 0


def y2_fk_iskonto_skor(fk_isk):
    """Y2: F/K İskontosu vs 3Y Ort (ANA YÖNTEM)"""
    if fk_isk is None: return 0
    if fk_isk > 50: return 15
    if fk_isk > 30: return 11
    if fk_isk > 15: return 6
    return 0


def y3_hedef_pot_skor(potansiyel):
    """Y3: Net Kâr Kapitalizasyonu Potansiyeli"""
    if potansiyel is None: return 0
    if potansiyel > 50: return 5
    if potansiyel > 25: return 3
    if potansiyel > 10: return 2
    return 0


def y4_pddd_skor(pddd):
    """Y4: PD/DD Değerleme"""
    if pddd is None: return 0
    if pddd < 1.0: return 25
    if pddd < 1.5: return 20
    if pddd < 2.0: return 15
    if pddd < 3.0: return 8
    return 0


def y5_roe_skor(roe):
    """Y5: ROE Kalite"""
    if roe is None: return 0
    if roe > 40: return 15
    if roe > 25: return 12
    if roe > 15: return 8
    if roe > 10: return 5
    return 0


def y6_momentum_skor(fk_ttm, fk_fwd):
    """Y6: Forward vs TTM Momentum"""
    if fk_ttm is None or fk_fwd is None or fk_ttm <= 0 or fk_fwd <= 0:
        return 0
    daralma = (fk_ttm - fk_fwd) / fk_ttm * 100
    if daralma > 50: return 10
    if daralma > 25: return 7
    if daralma > 0: return 4
    return 0


def forward_fk_skor(fwd_fk, sektor_adil):
    """Forward F/K vs Sektör Adil F/K"""
    if fwd_fk is None or sektor_adil is None or sektor_adil <= 0:
        return 0
    ratio = fwd_fk / sektor_adil
    if ratio < 0.5: return 25
    if ratio < 1.0: return 20
    if ratio < 1.5: return 15
    if ratio < 2.0: return 8
    return 0


def kar_surprizi_skor(qoq, yoy):
    """Kâr Sürprizi (QoQ+YoY)"""
    qoq_pos = qoq is not None and qoq > 0
    yoy_pos = yoy is not None and yoy > 0
    if qoq_pos and yoy_pos: return 15
    if qoq_pos or yoy_pos: return 10
    if qoq is not None and abs(qoq) < 10: return 3
    return 0


# ─── SİNYAL BELİRLEME ──────────────────────────────────────────────

def sinyal_belirle(skor, potansiyel):
    pot = potansiyel if potansiyel is not None else 0
    if skor >= 80 and pot > 50:
        return "🟢 ALTIN FIRSAT"
    if skor >= 80:
        return "🟢 DERİN İSKONTO"
    if skor >= 60:
        return "🟡 FORWARD UCUZ"
    if skor >= 40:
        return "🟠 TAVSİYE"
    if skor >= 30:
        return "⚪ İZLEME"
    return "🔴 ELEN"


# ─── ANA PİPELİNE ──────────────────────────────────────────────────

def tsv_oku(path):
    """TSV dosyasını dict listesi olarak oku"""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            rows.append(row)
    return rows


def pipeline(master_path, hist_path, faiz=42.5, faiz_trendi="indirim", donem="Q4_2025"):
    """Ana pipeline: master + hist → skorlanmış liste"""

    master = tsv_oku(master_path)
    hist = tsv_oku(hist_path)

    # Tarihsel çarpanları dict'e çevir
    hist_dict = {}
    for h in hist:
        kod = h.get("kod", h.get("hisse_senedi_kodu", ""))
        hist_dict[kod] = h

    # Faiz hesapla
    baz = adil_fk(faiz, faiz_trendi)
    print(f"⚙️  TCMB Faiz: %{faiz} ({faiz_trendi}) → Baz Adil F/K: {baz}x")

    sonuclar = []

    for row in master:
        kod = row.get("kod", "")
        sektor = row.get("sektor", "Diğer")

        # GYO/Holding → ayrı (basit skor)
        if sektor in ("Gayrimenkul", "Holding"):
            # Sadece PD/DD bazlı basit skor
            pddd = safe_float(row.get("pddd"))
            fk = safe_float(row.get("fk_ttm"))
            hist_row = hist_dict.get(kod, {})
            ort_pddd = safe_float(hist_row.get("ort_pddd_3y"))
            ort_fk = safe_float(hist_row.get("ort_fk_3y"))

            pddd_isk = iskonto_pct(pddd, ort_pddd)
            fk_isk = iskonto_pct(fk, ort_fk)

            # Çift filtre
            cift_ucuz = (pddd_isk is not None and pddd_isk > 0 and
                         fk_isk is not None and fk_isk > 0)
            skor = 0
            if pddd_isk and pddd_isk > 50: skor += 35
            elif pddd_isk and pddd_isk > 30: skor += 28
            elif pddd_isk and pddd_isk > 15: skor += 18
            else: skor += 5

            if fk_isk and fk_isk > 30: skor += 20
            elif fk_isk and fk_isk > 0: skor += 12

            if cift_ucuz: skor += 15

            pot = pddd_isk if pddd_isk else 0
            sinyal = sinyal_belirle(skor, pot)

            sonuclar.append({
                "kod": kod, "sektor": sektor, "total_skor": round(skor, 1),
                "potansiyel": round(pot, 1), "fk_ttm": fk, "fk_fwd": None,
                "sinyal": sinyal, "aciklama": f"Çift filtre: {'✅' if cift_ucuz else '❌'}",
                "y1": 0, "y2": 0, "y3": y4_pddd_skor(pddd), "y4": 0, "y5": 0, "y6": 0,
            })
            continue

        # Standart şirket
        s_adil = sektor_adil_fk(baz, sektor)
        if s_adil is None:
            s_adil = baz

        fk_ttm = safe_float(row.get("fk_ttm"))
        pddd = safe_float(row.get("pddd"))
        fd_favok = safe_float(row.get("fd_favok"))
        roe = safe_float(row.get("roe"))
        fk_fwd = safe_float(row.get("fk_fwd", row.get("fk_2025e")))
        nk_ttm = safe_float(row.get("nk_ttm", row.get("net_kar")))
        nk_qoq = safe_float(row.get("nk_qoq"))
        nk_yoy = safe_float(row.get("nk_yoy"))
        net_borc_pd = safe_float(row.get("net_borc_pd"))

        # Tarihsel
        hist_row = hist_dict.get(kod, {})
        ort_fk = safe_float(hist_row.get("ort_fk_3y"))
        ort_pddd = safe_float(hist_row.get("ort_pddd_3y"))
        ort_fd_favok = safe_float(hist_row.get("ort_fd_favok_3y"))

        # İskontolar
        fk_isk = iskonto_pct(fk_ttm, ort_fk)
        fd_favok_isk = iskonto_pct(fd_favok, ort_fd_favok)

        # Forward vs sektör adil
        fwd_skor = forward_fk_skor(fk_fwd, s_adil)

        # 6 Yöntem skorları
        s_y1 = y1_fd_favok_skor(fd_favok_isk)
        s_y2 = y2_fk_iskonto_skor(fk_isk)
        s_y3 = y3_hedef_pot_skor(fk_isk)  # proxy
        s_y4 = y4_pddd_skor(pddd)
        s_y5 = y5_roe_skor(roe)
        s_y6 = y6_momentum_skor(fk_ttm, fk_fwd)

        # Kâr sürprizi
        s_surpriz = kar_surprizi_skor(nk_qoq, nk_yoy)

        total = fwd_skor + s_y1 + s_y2 + s_surpriz + s_y4 + s_y5 + s_y6

        # Bonus / Ceza
        bonus = 0
        if pddd is not None and pddd < 1.0: bonus += 3
        elif pddd is not None and pddd < 1.5: bonus += 2

        if roe is not None and roe > 15: bonus += 2

        # v2.4 ceza: NK negatif
        if nk_ttm is not None and nk_ttm <= 0:
            total = max(total - 15, 0)

        # v2.4: Fwd F/K N/A → potansiyel ceza
        if fk_fwd is None:
            fk_isk = (fk_isk or 0) * 0.7

        total = round(total + bonus, 1)
        potansiyel = fk_isk if fk_isk else 0

        sinyal = sinyal_belirle(total, potansiyel)

        # v2.4: skor < 30 → ELEN
        if total < 30:
            sinyal = "🔴 ELEN"

        sonuclar.append({
            "kod": kod, "sektor": sektor, "total_skor": total,
            "potansiyel": round(potansiyel, 1),
            "fk_ttm": fk_ttm, "fk_fwd": fk_fwd,
            "sinyal": sinyal, "aciklama": "",
            "y1": s_y1, "y2": s_y2, "y3": s_y3, "y4": s_y4, "y5": s_y5, "y6": s_y6,
        })

    # Sırala
    sonuclar.sort(key=lambda x: x["total_skor"], reverse=True)

    print(f"📊 Toplam: {len(sonuclar)} hisse skorlandı")
    print(f"🟢 DERİN İSKONTO: {sum(1 for s in sonuclar if 'DERİN' in s['sinyal'])}")
    print(f"🟡 FORWARD UCUZ:  {sum(1 for s in sonuclar if 'FORWARD' in s['sinyal'])}")
    print(f"🔴 ELEN:          {sum(1 for s in sonuclar if 'ELEN' in s['sinyal'])}")

    return sonuclar


# ─── EXCEL ÇIKTI ──────────────────────────────────────────────────

SINYAL_RENK = {
    "🟢 ALTIN FIRSAT": "00FF00",
    "🟢 DERİN İSKONTO": "90EE90",
    "🟡 FORWARD UCUZ": "FFFF00",
    "🟠 TAVSİYE": "FFA500",
    "⚪ İZLEME": "D3D3D3",
    "🔴 ELEN": "FF6347",
}


def excel_yaz(sonuclar, cikti_path, donem="Q4_2025"):
    if not HAS_OPENPYXL:
        print("❌ openpyxl gerekli: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarama Sonuçları"

    # Başlıklar
    basliklar = ["Kod", "Sektör", "Skor", "Y1_FK", "Y2_PDDD", "Y3_FD", "Y4_ROE",
                 "Y5_Kaldıraç", "Y6_Mom", "Potansiyel_%", "FK_TTM", "FK_Fwd", "Sinyal"]
    for i, b in enumerate(basliklar, 1):
        c = ws.cell(row=1, column=i, value=b)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F4F4F")
        c.alignment = Alignment(horizontal="center")

    # Veri
    for idx, s in enumerate(sonuclar, 2):
        ws.cell(row=idx, column=1, value=s["kod"])
        ws.cell(row=idx, column=2, value=s["sektor"])
        ws.cell(row=idx, column=3, value=s["total_skor"])
        ws.cell(row=idx, column=4, value=s["y1"])
        ws.cell(row=idx, column=5, value=s["y2"])
        ws.cell(row=idx, column=6, value=s["y3"])
        ws.cell(row=idx, column=7, value=s["y4"])
        ws.cell(row=idx, column=8, value=s["y5"])
        ws.cell(row=idx, column=9, value=s["y6"])
        ws.cell(row=idx, column=10, value=s["potansiyel"])
        ws.cell(row=idx, column=11, value=s["fk_ttm"])
        ws.cell(row=idx, column=12, value=s["fk_fwd"])
        ws.cell(row=idx, column=13, value=s["sinyal"])

        # Renklendirme
        renk = SINYAL_RENK.get(s["sinyal"], "FFFFFF")
        for col in range(1, 14):
            ws.cell(row=idx, column=col).fill = PatternFill("solid", fgColor=renk)

    # Kolon genişlikleri
    for i in range(1, 14):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Dashboard sayfası
    ds = wb.create_sheet("Dashboard")
    ds.cell(row=1, column=1, value=f"BIST Tarama {donem}").font = Font(bold=True, size=16)
    ds.cell(row=2, column=1, value=f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ds.cell(row=3, column=1, value=f"Toplam: {len(sonuclar)} hisse")

    row = 5
    for sinyal, renk in SINYAL_RENK.items():
        sayi = sum(1 for s in sonuclar if s["sinyal"] == sinyal)
        ds.cell(row=row, column=1, value=sinyal)
        ds.cell(row=row, column=2, value=sayi)
        ds.cell(row=row, column=1).fill = PatternFill("solid", fgColor=renk)
        row += 1

    wb.save(cikti_path)
    print(f"✅ Excel kaydedildi: {cikti_path}")


# ─── TSV ÇIKTI ──────────────────────────────────────────────────

def tsv_yaz(sonuclar, cikti_path):
    with open(cikti_path, "w", encoding="utf-8") as f:
        f.write("kod\tsektor\ttotal_skor\tpotential\tfk_ttm\tfk_2025e\tsinyal\n")
        for s in sonuclar:
            f.write(f"{s['kod']}\t{s['sektor']}\t{s['total_skor']}\t"
                    f"{s['potansiyel']}\t{s.get('fk_ttm','')}\t"
                    f"{s.get('fk_fwd','')}\t{s['sinyal']}\n")
    print(f"✅ TSV kaydedildi: {cikti_path}")


# ─── CLI ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="BIST Tarama Pipeline v2.4")
    parser.add_argument("master", help="master.tsv dosya yolu")
    parser.add_argument("hist", help="hist_master.tsv dosya yolu")
    parser.add_argument("--faiz", type=float, default=42.5, help="TCMB politika faizi (%%)")
    parser.add_argument("--trend", default="indirim", choices=["indirim","sabit","artirim"])
    parser.add_argument("--donem", default="Q4_2025", help="Bilanço dönemi")
    parser.add_argument("--output", default=None, help="Çıktı dosya adı (xlsx)")
    args = parser.parse_args()

    sonuclar = pipeline(args.master, args.hist, args.faiz, args.trend, args.donem)

    out_name = args.output or f"BIST_{args.donem}_6YONTEM_TARAMA"
    excel_yaz(sonuclar, f"{out_name}.xlsx", args.donem)
    tsv_yaz(sonuclar, f"{out_name}.tsv")


if __name__ == "__main__":
    main()
